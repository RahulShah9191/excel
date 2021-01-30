import pandas as pd
from openpyxl import Workbook, load_workbook
import errno
from utils import file_handler, string_lib





class XlsxLib:

    def __init__(self, path_to_file=None, sheet_name=None, read_only=False, header=0,
                 create_if_not_exists=False, overwrite_if_exists=False):
        """

        :param path_to_file: If absolute path to excel file is sent, then we create a workbook instance with that file
                            and use below parameters for further handling.
        :param sheet_name: If sheet_name is provided, then create a sheet with the provided name as create_if_not_exists
        :param read_only: Default=False; If set to True, then a read_only file is created!
        :param create_if_not_exists: Default=False; If set to True, then a new file is created if not present.
        :param overwrite_if_exists: Default=False; If set to True, it allows to overwrite file if exists, while creation
        Usage: var = XlsxLib(workbook=already_created)
        """
        #global LOGGER
        #LOGGER = None
        self.path_to_file = path_to_file
        self.sheet_name = sheet_name
        self.read_only = read_only
        self.header = header
        self.create_if_not_exists = create_if_not_exists
        self.overwrite_if_exists = overwrite_if_exists
        self.workbook = None
        self.status = True

        if not string_lib.is_string_empty_or_none(path_to_file):
            try:
                if os.path.dirname(path_to_file) == "" or os.path.dirname(path_to_file) is None:
                    print("Absolute path to file - {f}, has not been provided!".format(f=path_to_file))
                elif file_handler.is_file_present(path_to_file) and os.path.dirname(path_to_file) is not None:
                    self.status = self.load_workbook_from_excel_file()
                elif create_if_not_exists and not file_handler.is_file_present(path_to_file):
                    self.status, self.workbook = XlsxLib.create_new_excel_file(path_to_new_file=self.path_to_file,
                                                                               sheet_name=self.sheet_name,
                                                                               sheet_index=None,
                                                                               overwrite_if_exists=self.overwrite_if_exists)
                else:
                    self.status = False
                    print("{f} - File does not exist!".format(f=self.path_to_file))
            except Exception as e:
                print("Unable to access or create file- {f}! \nException: {e}".format(f=self.path_to_file, e=e))
                self.status = False
        else:
            print("{f} - File path can not be empty/None!".format(f=self.path_to_file))
            self.status = False
            raise FileNotFoundError(errno.ENOENT, os.strerror(errno.ENOENT), path_to_file)

    def load_workbook_from_excel_file(self):
        status = False
        try:
            if file_handler.is_file_present(self.path_to_file):
                self.workbook = load_workbook(filename=self.path_to_file, read_only=self.read_only)
                status = True
            else:
                print("File not present or sheet_name not present in file!")
                status = False
        except Exception as e:
            print("Unable to load workbook from excel file {f}".format(f=self.path_to_file))

        return status

    @staticmethod
    def save_excel_file(path_to_file, workbook):
        status = False
        try:
            workbook.save(path_to_file)
            file_handler.change_file_permissions(path_to_file, permissions="777")
            status = True
        except Exception as e:
            LOGGER.critical("Unable to save excel file :: " + path_to_file + " \n Exception :: {0}".format(e))
            status = False
        return status

    @staticmethod
    def create_new_excel_file(path_to_new_file, sheet_name=None, sheet_index=None, overwrite_if_exists=False):
        create_file = False
        workbook = None
        status = False
        try:
            if file_handler.is_file_present(path_to_new_file):
                if overwrite_if_exists:
                    create_file = True
                    LOGGER.warning("Overwriting file, since overwrite_if_exists = {o}".format(o=overwrite_if_exists))
                else:
                    create_file = False
                    workbook = load_workbook(path_to_new_file)
                    print("File {f}, already exists and overwrite is False!".format(f=path_to_new_file))
            else:
                create_file = True

            if create_file:
                workbook = Workbook()
                try:
                    workbook.create_sheet(title=sheet_name, index=sheet_index)
                    status = True
                except Exception as e:
                    print(
                        "Unable to create file {f} with sheet {s} \n Exception :: ".format(
                                                                                s=sheet_name, f=path_to_new_file, e=e))
                    status = False
                if status:
                    XlsxLib.save_excel_file(path_to_new_file, workbook)
                    print("Created a new excel file :: " + path_to_new_file)
                    status = True
            else:
                status = False

        except Exception as e:
            print("Unable to create file {f} with sheet {s} \n Exception :: ".format(
                                                                                s=sheet_name, f=path_to_new_file, e=e))
            status = False

        return status, workbook

    def check_if_sheet_present_in_excel_file(self, sheet_name):
        status = None
        try:
            ws = self.workbook[sheet_name]
            status = True
        except Exception as e:
            print("Unable to load workbook!")
            status = False
        finally:
            return status

    def remove_sheet_from_excel_file(self, sheet_name):
        if self.check_if_sheet_present_in_excel_file(sheet_name=sheet_name):
            try:
                self.workbook.remove(self.workbook[sheet_name])
                status = True if XlsxLib.save_excel_file(self.path_to_file, self.workbook) else False
            except Exception as e:
                print("Unable to remove sheet from excel file! \n Exception :: {0}".format(e))
                status = False
        else:
            print("Sheet {s}, not present in file {f}!".format(s=sheet_name, f=self.path_to_file))
            status = False
        self.status = status
        return status

    def create_new_sheet_in_excel_file(self, sheet_name, sheet_index=None, create_excel_if_not_exists=True):
        status = False
        create_sheet = False
        if file_handler.is_file_present(self.path_to_file):
            create_sheet = True
        elif file_handler.is_file_present(self.path_to_file) is False and create_excel_if_not_exists is True:
            print(" create_excel_if_not_exists = {c} | Trying to create a new excel file at {f}".format(
                                                                    c=create_excel_if_not_exists, f=self.path_to_file))
            if XlsxLib.create_new_excel_file(path_to_new_file=self.path_to_file,
                                             sheet_name=sheet_name,
                                             sheet_index=sheet_index):
                # The if condition returns True when we successfully create a excel file with sheet_name
                create_sheet = False
                status = True
            else:
                create_sheet = True
        else:
            create_sheet = False

        if create_sheet and not status:
            try:
                self.workbook = load_workbook(self.path_to_file)
                is_sheet_present = self.check_if_sheet_present_in_excel_file(sheet_name)
                if not is_sheet_present:
                    self.workbook.create_sheet(sheet_name)
                    XlsxLib.save_excel_file(path_to_file=self.path_to_file, workbook=self.workbook)
                    status = True
                else:
                    LOGGER.warning("Sheet with name {s} already exists!".format(s=sheet_name))
                    status = False
            except Exception as e:
                print("Unable to create sheet in workbook! \n Exception ::  {0}".format(e))
                status = False
        else:
            print("Unable to create sheet in excel :: {f}".format(f=self.path_to_file))
            status = False

        self.status = status
        return status

    def read_data_by_cell_coordinates(self, sheet_name, row_num, column_num, read_only=True):
        """

        :param path_to_file:
        :param sheet_name:
        :param row_num:
        :param column_num:
        :param read_only:
        :param workbook:
        :return:
        ## Add support for if sheet not present in file in the workbook passed, then what to do, currently raising exception
        """
        if self.check_if_sheet_present_in_excel_file(sheet_name):
            try:
                ws = self.workbook[sheet_name]
                data = ws.cell(row=int(row_num), column=int(column_num)).value
                if str(data).strip() == "None":
                    return ""
                else:
                    return str(data).strip()
            except Exception as e:
                print("Unable to read data from excel :: {f} \n Exception :: {e}".format(
                                                                                            f=self.path_to_file, e=e))
                return False
        else:
            LOGGER.warning("Sheet does not exists!")
            return False

    def write_data_by_cell_coordinates(self, sheet_name, row_num, column_num, data):
        """

        :param sheet_name:
        :param row_num:
        :param column_num:
        :param data:
        :param workbook:
        :return:
        ## Add support for if sheet not present in file in the workbook passed, then what to do, currently raising exception
        """
        if self.check_if_sheet_present_in_excel_file(sheet_name):
            try:
                ws = self.workbook[sheet_name]
                ws.cell(row=int(row_num), column=int(column_num)).value = data
                XlsxLib.save_excel_file(self.path_to_file, self.workbook)
                self.status = True
                return True
            except Exception as e:
                print("Unable to write data to file {f}\n Exception :: {e}".format(f=self.path_to_file, e=e))
                self.status = False
                return False
        else:
            LOGGER.warning("Sheet {s} does not exists in {f}!".format(s=sheet_name, f=self.path_to_file))
            self.status = False
            return False

    def get_max_row_count_from_excel_sheet(self, sheet_name):
        """

        :param sheet_name:
        :param workbook:
        :return:
        ## Add support for if sheet not present in file in the workbook passed, then what to do, currently raising exception
        """
        result = False
        if self.check_if_sheet_present_in_excel_file(sheet_name=sheet_name):
            try:
                #print(self.workbook.sheetnames)
                _ws = self.workbook[sheet_name]
                row_count = _ws.max_row
                #print("row_count=", row_count)
                result = row_count
            except Exception as e:
                print("Unable to get max row count from sheet {s} in file {f} \nException :: {e}".format(
                    s=sheet_name, f=self.path_to_file, e=e))
                result = False
            finally:
                return result
        else:
            LOGGER.warning("Sheet {s} does not exists in file {f}!".format(s=sheet_name, f=self.path_to_file))
            result = False
            return result

    def get_max_column_count_from_excel_sheet(self, sheet_name):
        """

        :param path_to_file:
        :param sheet_name:
        :param workbook:
        :return:
        ## Add support for if sheet not present in file in the workbook passed, then what to do, currently raising exception
        """
        result = False
        if sheet_name is not None and sheet_name.strip(" ") != "":
            if sheet_name in self.workbook.sheetnames:
                try:
                    #print(self.workbook.sheetnames)
                    ws = self.workbook[sheet_name]
                    col_count = ws.max_column + 1
                    #print("col_count=", col_count)
                    result = col_count
                except Exception as e:
                    print("Unable to get max column count from sheet {s} in file {f} \nException :: {e}".format(
                        s=sheet_name, f=self.path_to_file, e=e))
                    result = False
            else:
                LOGGER.warning("Sheet {s} does not exists in file {f}!".format(s=sheet_name, f=self.path_to_file))
                result = False
        else:
            print("sheet_name can not be empty or None!")
            result = False
        return result

    def append_to_excel_sheet_row(self, sheet_name, data, row_num=None, overwrite_data=False):
        """

        :param sheet_name:
        :param data:
        :param row_num:
        :param overwrite_data:
        :param workbook:
        :return:
        """
        status = False
        if self.check_if_sheet_present_in_excel_file(sheet_name=sheet_name):
            status = True
        else:
            print("Sheet - {s} not present in file - {f}!".format(s=sheet_name, f=self.path_to_file))
            status = False
        try:
            ws = self.workbook[sheet_name]
            row_count = ws.max_row + 1
            col_count = ws.max_column + 1

            data_error_msg = "Data can be either a list of list string/int, representing a row, i.e. [[ 1, 2, 3]] OR "
            data_error_msg += "A plain string/int representing a single value, i.e 1 or 'PubMatic'"
            overwrite_error_msg = "row_num can not be less than maximum row having data, when overwrite_data is False!"

            status = True if all(isinstance(el, list) for el in data) else True if isinstance(data, (str, int)) else False
            if status is False:
                print(data_error_msg)
                status = False
            else:
                row_count = ws.max_row + 1
                col_count = ws.max_column + 1

                if row_num is None and status:
                    if isinstance(data, list) and all(isinstance(el, list) for el in data):
                        for element in data:
                            ws.append(element)
                        status = True
                    elif isinstance(data, (str, int)) and status:
                        ws.cell(row=row_count, column=1).value = data
                        status = True
                    else:
                        print(data_error_msg)
                        status = False
                elif (row_num >= row_count) or (row_num < row_count and overwrite_data):
                    if isinstance(data, list) and all(isinstance(el, list) for el in data):
                        for element in data:
                            c = 1
                            for d in element:
                                ws.cell(row=row_num, column=c).value = d
                                c = c + 1
                            row_num = row_num + 1
                        status = True
                    elif isinstance(data, (str, int)) and status:
                        ws.cell(row=row_num, column=1).value = data
                        status = True
                    else:
                        print(data_error_msg)
                        status = False
                else:
                    print(overwrite_error_msg)
                    status = False
                status = XlsxLib.save_excel_file(self.path_to_file, self.workbook) if status else False
                if not status:
                    print("Unable to save file {f}\n Exception: {e}".format(f=self.path_to_file))
        except Exception as e:
            print("Unable to write data to file {f}\n Exception: {e}".format(f=self.path_to_file, e=e))
            status = False
        self.status = status
        return status

    def append_to_excel_sheet_column(self, sheet_name, data, column_num=None, overwrite_data=False):
        """

        :param sheet_name:
        :param data:
        :param column_num:
        :param overwrite_data:
        :param workbook:
        :return:
        """
        status = False
        if self.check_if_sheet_present_in_excel_file(sheet_name=sheet_name):
            status = True
        else:
            print("Sheet - {s} not present in file - {f}!".format(s=sheet_name, f=self.path_to_file))
            status = False
        try:
            ws = self.workbook[sheet_name]
            row_count = ws.max_row + 1
            col_count = ws.max_column + 1

            data_error_msg = "Data can be either a list of list string/int, representing a row, i.e. [[ 1, 2, 3]] OR "
            data_error_msg += "A plain string/int representing a single value, i.e 1 or 'PubMatic'"
            overwrite_error_msg = "row_num can not be less than maximum row having data, when overwrite_data is False!"

            status = True if all(isinstance(el, list) for el in data) else True if isinstance(data, (str, int)) else False
            if status is False:
                print(data_error_msg)
                status = False
            else:
                row_count = ws.max_row + 1
                col_count = ws.max_column + 1

                if column_num is None and status:
                    if isinstance(data, list) and all(isinstance(el, list) for el in data):
                        for element in data:
                            r = 1
                            for d in element:
                                ws.cell(row=r, column=col_count).value = d
                                r = r + 1
                            col_count = ws.max_column + 1
                        status = True
                    elif isinstance(data, (str, int)) and status:
                        ws.cell(row=row_count, column=1).value = data
                        status = True
                    else:
                        print(data_error_msg)
                        status = False
                elif (column_num >= col_count) or (column_num < col_count and overwrite_data):
                    if isinstance(data, list) and all(isinstance(el, list) for el in data):
                        for element in data:
                            r = 1
                            for d in element:
                                ws.cell(row=r, column=column_num).value = d
                                r = r + 1
                            column_num = column_num + 1
                        status = True
                    elif isinstance(data, (str, int)) and status:
                        ws.cell(row=column_num, column=1).value = data
                        status = True
                    else:
                        print(data_error_msg)
                        status = False
                else:
                    print(overwrite_error_msg)
                    status = False
                status = XlsxLib.save_excel_file(self.path_to_file, self.workbook) if status else False
                if not status:
                    print("Unable to save file {f}".format(f=self.path_to_file))
        except Exception as e:
            print("Unable to write data to file {f}\n Exception: {e}".format(f=self.path_to_file, e=e))
            status = False
        self.status = status
        return status

    @staticmethod
    def copy_range_from_excel_sheet(path_to_file, sheet_name, start_row, start_col, end_row, end_col):
        status = False
        try:
            if file_handler.is_file_present(path_to_file) and path_to_file is not None:
                workbook = load_workbook(path_to_file)
                status = True
            else:
                print("File not present!!")
                status = False
        except Exception as e:
            exception_msg = "Unable to load {f}! \nException: {e}".format(f=path_to_file, e=e)
            print(exception_msg)
            #raise Exception("Unable to load {f}! \nException: {e}".format(f=path_to_file, e=e))
        if sheet_name is not None and sheet_name != "".strip(" "):
            if sheet_name in workbook.sheetnames and status:
                try:
                    ws = workbook[sheet_name]
                    range_selected = []
                    # Loops through selected Rows
                    for i in range(start_row, end_row + 1, 1):
                        # Appends the row to a row_selected list
                        row_selected = []
                        for j in range(start_col, end_col + 1, 1):
                            row_selected.append(ws.cell(row=i, column=j).value)
                        # Adds the row_selected List and nests inside the range_selected
                        range_selected.append(row_selected)
                    return range_selected
                except Exception as e:
                    print("Unable to find sheet {s} in file {f}!".format(s=sheet_name, f=path_to_file))
                    status = False
            else:
                print("Sheet {s}, not present in file {f}!".format(s=sheet_name, f=path_to_file))
                status = False
        else:
            print("Sheet name cannot be None or empty{s}!".format(s=sheet_name))
        return status

    @staticmethod
    def paste_range_to_excel_sheet(path_to_file, sheet_name, copied_data, start_row, start_col,
                                   end_row=None, end_col=None):
        status = False
        if file_handler.is_file_present(path_to_file):
            workbook = load_workbook(path_to_file)
            status = True
        else:
            print("File not present!!")
            status = False

        end_row = len(copied_data) if end_row is None and copied_data is not None else end_row
        end_col = max([len(element) for element in copied_data]) if end_col is None and copied_data is not None else end_col

        if sheet_name is not None and sheet_name != "".strip(" "):
            if sheet_name in workbook.sheetnames:
                try:
                    ws = workbook[sheet_name]
                    count_row = 0
                    for i in range(start_row, end_row + 1, 1):
                        count_col = 0
                        for j in range(start_col, end_col + 1, 1):
                            ws.cell(row=i, column=j).value = copied_data[count_row][count_col]
                            count_col += 1
                        count_row += 1
                    XlsxLib.save_excel_file(path_to_file, workbook)
                    status = True
                except Exception as e:
                    print("Unable to find the sheet {s} in file {f}!".format(s=sheet_name, f=path_to_file))
                    status = False
            else:
                print("Sheet {s}, not present in file {f}!".format(s=sheet_name, f=path_to_file))
                status = False
        else:
            print("Sheet name cannot be None or empty{s}!".format(s=sheet_name))
        return status

    @staticmethod
    def copy_and_paste_data_to_excel_sheet(path_to_src_file, src_sheet_name,
                                     src_start_row, src_start_col,
                                     src_end_row, src_end_col,
                                     tgt_path_to_file, tgt_sheet_name,
                                     tgt_start_row, tgt_start_col,
                                     tgt_end_row, tgt_end_col):
        status = False
        try:
            range_selected = XlsxLib.copy_range_from_excel_sheet(path_to_file=path_to_src_file, sheet_name=src_sheet_name,
                                                              start_col=src_start_col, start_row=src_start_row,
                                                              end_col=src_end_col, end_row=src_end_row)
            if range_selected is not None:
                pasting_result = XlsxLib.paste_range_to_excel_sheet(path_to_file=tgt_path_to_file,
                                                                    sheet_name=tgt_sheet_name,
                                                                    start_row=tgt_start_row,
                                                                    start_col=tgt_start_col,
                                                                    end_row=tgt_end_row,
                                                                    end_col=tgt_end_col,
                                                                    copied_data=range_selected)
                if pasting_result:
                    print("Copied and pasted successfully!")
                    status = True
                else:
                    print("Unable to paste data to target file!")
                    status = False
            else:
                print("No data copied! \n Kindly check the range for selecting data!")
                status = False
        except Exception as e:
            print("Exception occurred while copying & pasting data to file\n Exception :: {0}".format(e))
        return status

    def get_sheet_data_as_df(self, sheet_name, header=0):
        """

        :param sheet_name: sheet name in the excel sheet
        :param header: Default header=0, skip first row while returning data. If None, all rows data is sent back.
        :return: data frame of the data from sheet
        """
        if not string_lib.is_string_empty_or_none(sheet_name):
            try:
                sheet_data_df = pd.read_excel(self.path_to_file, sheet_name=sheet_name, header=header)
                return sheet_data_df
            except Exception as e:
                print("Unable to read sheet data to data frame!")
                return False
        else:
            print("Sheet name {s}, can not be empty!".format(s=sheet_name))
            return False

    def get_sheet_data_as_list(self, sheet_name, header=0):
        """

        :param sheet_name: sheet name in the excel sheet
        :param header: Default header=0, skip first row while returning data. If None, all rows data is sent back.
        :return: list of list, i.e list of rows.
        """
        if not string_lib.is_string_empty_or_none(sheet_name):
            try:
                all_rows_data = []
                for row in self.workbook[sheet_name].iter_rows():
                    all_rows_data.append([cell.value for cell in row])
                if header == 0:
                    all_rows_data.pop(0)
                return all_rows_data
            except Exception as e:
                print("Unable to read sheet data to list!\n Exception :: {e}".format(e=e))
        else:
            print("Sheet name {s}, can not be empty!".format(s=sheet_name))
            return False
